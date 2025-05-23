import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference
from openpyxl.formula.translate import Translator
import xlwings as xw
from transformers import AutoTokenizer, AutoModelForCausalLM
import torch
import shutil
from fastapi import FastAPI
from pydantic import BaseModel

app = FastAPI()


# 🔄 Load Hinglish-understanding AI model
print("🔄 Loading phi-1_5 model...")
tokenizer = AutoTokenizer.from_pretrained("microsoft/phi-1_5")
model = AutoModelForCausalLM.from_pretrained("microsoft/phi-1_5")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)
print("✅ Model loaded")

# 🧠 Generate Excel formula
def generate_formula(instruction):
    prompt = f"Generate only Excel formula for: {instruction}\nFormula:"
    inputs = tokenizer(prompt, return_tensors="pt").to(device)
    output = model.generate(**inputs, max_length=100)
    result = tokenizer.decode(output[0], skip_special_tokens=True)
    return result.replace("Output:", "").replace("Formula:", "").strip()

# 📊 Detect chart instruction
def is_chart_instruction(cmd):
    return any(word in cmd.lower() for word in ["chart", "plot", "bar", "graph"])

# 📈 Create bar chart
def create_chart(ws, min_col, max_col, min_row, max_row):
    chart = BarChart()
    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, f"{chr(65 + max_col + 1)}5")

# 📌 Conditional formatting
def apply_conditional_formatting(ws, col_letter, start_row, end_row):
    for row in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{row}"]
        try:
            if isinstance(cell.value, (int, float)) and cell.value > 100:
                cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
        except:
            continue

# 📂 Detect open Excel workbook path
def get_open_excel_path():
    try:
        for wb in xw.apps.active.books:
            if wb.name.endswith(".xlsx") or wb.name.endswith(".xlsm"):
                return wb.fullname
    except Exception:
        return None

# 🧠 Apply formula to all rows
def apply_formula_to_sheet(ws, formula, start_row, target_col, max_row):
    ws.cell(row=1, column=target_col, value="Result")
    for i in range(start_row, max_row + 1):
        translated = Translator(formula, origin="B2").translate_formula(f"{chr(65 + target_col - 2)}{i}")
        ws.cell(row=i, column=target_col, value=f"={translated}")

# 🧠 Apply pivot table logic (placeholder)
def apply_pivot_logic(ws):
    print("ℹ️ Pivot Table: AI assisted automation will require VBA or user-defined range setup")

# 🧠 Apply macro (as string insertion)
def insert_macro_code():
    return """
Sub AutoFilterColumnA()
    Columns("A:A").AutoFilter
End Sub
"""

# 🧠 Main logic
def process_live_excel():
    path = get_open_excel_path()
    if not path:
        print("❌ कोई Excel फाइल खुली नहीं है।")
        return

    # ✅ Backup before overwrite
    if path.endswith(".xlsx"):
        backup_path = path.replace(".xlsx", "_backup.xlsx")
        shutil.copyfile(path, backup_path)
        print(f"🧾 Backup saved as: {backup_path}")

    df = pd.read_excel(path, sheet_name=None)
    wb = load_workbook(path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = df[sheet_name]
        total_rows = len(data)
        total_cols = len(data.columns)

        print(f"📄 Processing sheet: {sheet_name} ({total_rows} rows)")
        cmd = input("🧠 क्या करना है? (हिंग्लिश में लिखें, 'exit' टाइप करें बंद करने के लिए): ")
        if cmd.lower() in ["exit", "quit"]:
            print("👋 बंद किया गया।")
            break

        if is_chart_instruction(cmd):
            create_chart(ws, min_col=2, max_col=total_cols, min_row=1, max_row=total_rows + 1)
            print("📊 Chart created.")

        elif "pivot" in cmd.lower():
            apply_pivot_logic(ws)

        elif "macro" in cmd.lower():
            if not path.endswith(".xlsm"):
                print("⚠️ Please save the file as .xlsm for macro support.")
            else:
                vba_code = insert_macro_code()
                with open("macro_module.bas", "w") as f:
                    f.write(vba_code)
                print("🧩 Macro code saved in macro_module.bas. Please import manually into Excel.")

        elif "highlight" in cmd.lower():
            apply_conditional_formatting(ws, col_letter="B", start_row=2, end_row=total_rows + 1)
            print("✨ Conditional formatting applied.")

        else:
            formula = generate_formula(cmd)
            print(f"📥 Formula Generated: {formula}")
            if formula:
                apply_formula_to_sheet(ws, formula, start_row=2, target_col=total_cols + 1, max_row=total_rows + 1)
                print(f"✅ Formula applied on sheet: {sheet_name}")

    wb.save(path)
    print(f"💾 All changes saved to: {path}")

if __name__ == "__main__":
    process_live_excel()

class Command(BaseModel):
    instruction: str

@app.post("/ai-command")
def handle_ai_command(cmd: Command):
    formula = generate_formula(cmd.instruction)
    return {"output": formula}
